import os
import sqlite3

import pandas as pd
from aiogram import Bot, Router, F
from aiogram.filters import Command, or_f
from dotenv import load_dotenv
from sqlalchemy.ext.asyncio import AsyncSession
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from aiogram.types import FSInputFile, Message

from database.orm_database import orm_add_house, orm_add_status, orm_get_start_id_status, \
    orm_get_start_id_life, orm_set_status_sold, orm_get_start_1, orm_get_start_12, orm_get_start_id_people
from filter.chat_filtre import ChatTypeFilter, IsAdmin
from handlers.user.user_handlers import user_router
from kb.inline import inline_house_1, inline_object_type, inline_number_floors, inline_wall_material, inline_yes, \
    inline_yes_no, inline_post
from kb.replay import admin_menu

load_dotenv()

admin_router = Router()

admin_router.message.filter(ChatTypeFilter(['private']), IsAdmin())

admin_router = Router()


class House(StatesGroup):
    """
    Состояния для поэтапного добавления объекта недвижимости (дома).
    Каждый шаг собирает отдельный параметр дома.
    """
    step_1 = State()  # Материал стен
    step_2 = State()  # Водопровод (да/нет)
    step_3 = State()  # Природный газ (да/нет)
    step_4 = State()  # Электричество (да/нет)
    step_5 = State()  # Наличие скважины (да/нет)
    step_6 = State()  # Наличие бани (да/нет)
    step_7 = State()  # Наличие гаража (да/нет)
    step_8 = State()  # Площадь дома (текст)
    step_9 = State()  # Площадь участка (текст)
    step_10 = State()  # Цена (текст)
    step_11 = State()  # Описание (текст)
    step_12 = State()  # Фото объекта
    step_13 = State()


@admin_router.message(Command('admin'))
async def admin_start(message: Message, state: FSMContext):
    """
    Обработка команды /admin.
    Очищает состояние FSM и отправляет главное меню администратора.
    """
    await state.clear()
    await message.answer('Вот ваше меню', reply_markup=admin_menu)


# @admin_router.message(F.text == 'Добавить объект')
# async def start_house(message: Message):
#     """
#     Начинает процесс добавления объекта недвижимости.
#     Отправляет пользователю выбор категории объекта.
#     """
#     await message.answer('Выберите категорию', reply_markup=await inline_house_1())


@admin_router.callback_query(F.data == 'продать_дом')
async def sell_house(callback: CallbackQuery):
    """
    Обработка выбора "продать дом" - выводит список типов объектов.
    """
    await callback.answer()
    await callback.message.edit_text('Вид объекта', reply_markup=await inline_object_type())


@admin_router.callback_query(or_f(F.data == 'Дом', F.data == 'Дача', F.data == 'Котедж'))
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Сохраняет выбранный тип объекта и предлагает выбрать количество этажей.
    """
    await callback.answer()
    await state.update_data(type_object=callback.data)
    await callback.message.edit_text('Количество этажей', reply_markup=await inline_number_floors())


@admin_router.callback_query(or_f(F.data == '1', F.data == '2', F.data == '3', F.data == '4', F.data == '5'))
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Сохраняет количество этажей, спрашивает материал стен и переводит FSM в step_1.
    """
    await callback.answer()
    await state.update_data(number_floors=callback.data)
    await callback.message.edit_text('Материал стен', reply_markup=await inline_wall_material())
    await state.set_state(House.step_1)


@user_router.callback_query(House.step_1)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 1: Получение материала стен.
    Спрашивает про водопровод, переводит FSM в step_2.
    """
    await callback.answer()
    await state.update_data(wall_material=callback.data)
    await callback.message.edit_text('Водопровод', reply_markup=await inline_yes())
    await state.set_state(House.step_2)


@user_router.callback_query(House.step_2)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 2: Водоснабжение (да/нет).
    Спрашивает про природный газ, переводит FSM в step_3.
    """
    await callback.answer()
    await state.update_data(water_supply=callback.data)
    await callback.message.edit_text('Природный газ', reply_markup=await inline_yes())
    await state.set_state(House.step_3)


@user_router.callback_query(House.step_3)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 3: Природный газ.
    Спрашивает про электричество, переводит FSM в step_4.
    """
    await callback.answer()
    await state.update_data(natural_gas=callback.data)
    await callback.message.edit_text('Электричество', reply_markup=await inline_yes())
    await state.set_state(House.step_4)


@user_router.callback_query(House.step_4)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 4: Электричество.
    Спрашивает про наличие скважины, переводит FSM в step_5.
    """
    await callback.answer()
    await state.update_data(electricity=callback.data)
    await callback.message.edit_text('Скважина', reply_markup=await inline_yes_no())
    await state.set_state(House.step_5)


@user_router.callback_query(House.step_5)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 5: Скважина.
    Спрашивает про баню, переводит FSM в step_6.
    """
    await callback.answer()
    await state.update_data(well=callback.data)
    await callback.message.edit_text('Баня', reply_markup=await inline_yes_no())
    await state.set_state(House.step_6)


@user_router.callback_query(House.step_6)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 6: Баня.
    Спрашивает про гараж, переводит FSM в step_7.
    """
    await callback.answer()
    await state.update_data(bathhouse=callback.data)
    await callback.message.edit_text('Гараж', reply_markup=await inline_yes_no())
    await state.set_state(House.step_7)


@user_router.callback_query(House.step_7)
async def sell_house(callback: CallbackQuery, state: FSMContext):
    """
    Шаг 7: Гараж.
    Спрашивает площадь дома, переводит FSM в step_8.
    """
    await callback.answer()
    await state.update_data(garag=callback.data)
    await callback.message.edit_text('Площадь дома м²:', reply_markup=None)
    await state.set_state(House.step_8)


@user_router.message(House.step_8, F.text)
async def sell_house(message: Message, state: FSMContext):
    """
    Шаг 8: Получение площади дома (текст).
    Запрашивает площадь участка, переводит FSM в step_9.
    """
    await state.update_data(square=message.text)
    await message.answer('Площадь участка в сотках:')
    await state.set_state(House.step_9)


@user_router.message(House.step_9, F.text)
async def sell_house(message: Message, state: FSMContext):
    """
    Шаг 9: Получение площади участка.
    Запрашивает цену, переводит FSM в step_10.
    """
    await state.update_data(plot_area=message.text)
    await message.answer('Цена:')
    await state.set_state(House.step_13)

@user_router.message(House.step_13, F.text)
async def sell_house(message: Message, state: FSMContext):

    await state.update_data(price=message.text)
    await message.answer('Местоположение:')
    await state.set_state(House.step_10)


@user_router.message(House.step_10, F.text)
async def sell_house(message: Message, state: FSMContext):
    """
    Шаг 10: Получение цены.
    Запрашивает описание, переводит FSM в step_11.
    """
    await state.update_data(geo=message.text)
    await message.answer('Описание:')
    await state.set_state(House.step_11)


@user_router.message(House.step_11, F.text)
async def sell_house(message: Message, state: FSMContext):
    """
    Шаг 11: Получение описания.
    Запрашивает фотографию, переводит FSM в step_12.
    """
    await state.update_data(text_house=message.text)
    await message.answer('Загрузите фотографию:')
    await state.set_state(House.step_12)


@user_router.message(House.step_12, F.photo)
async def sell_house(message: Message, state: FSMContext):
    """
    Шаг 12: Получение фотографии объекта.
    Формирует итоговое сообщение с деталями объекта и кнопками публикации.
    """
    await state.update_data(photo=message.photo[-1].file_id)
    await state.update_data(status='Продается')
    await state.update_data(fio_buy=None)
    data = await state.get_data()
    text = (   f"Вид объекта: {data.get('title')}\n"
            f"Местоположение: {data.get('address')}\n"
            f"Ремонт: {data.get('repair')}\n"
            f"Площадь дома м²: {data.get('area')}\n"
            f"Тип здания: {data.get('building_type')}\n"
            f"Цена: {data.get('price')}\n\n"
            f"Природный газ: {data.get('gas_supply')}")
    sent_message = await message.answer_photo(f'{message.photo[-1].file_id}', text, reply_markup=await inline_post())
    await state.update_data(original_message_id=sent_message.message_id)


@user_router.callback_query(F.data == 'Опубликовать')
async def publish_house(callback: CallbackQuery, state: FSMContext, session: AsyncSession, bot: Bot):
    """
    Обработка кнопки "Опубликовать".
    Сохраняет объект в базу, очищает состояние, уведомляет пользователя.
    """
    await callback.answer()
    try:
        data = await state.get_data()
        original_message_id = data.get("original_message_id")
        await bot.edit_message_reply_markup(
            chat_id=callback.from_user.id, message_id=original_message_id, reply_markup=None)
    except Exception:
        pass
    data = await state.get_data()
    id_status = await orm_add_house(session, data)
    status = {
        'id_house': id_status,
        'like': 0,
        'like_nick': None,
        'watch': 0,
    }
    await orm_add_status(session, status)
    await state.clear()
    await callback.message.answer('Успешно опубликовано')


@user_router.callback_query(F.data == 'Отмена')
async def cancel_house(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """
    Обработка кнопки "Отмена" при добавлении дома.
    Очищает состояние и возвращает в главное меню.
    """
    await callback.answer()
    try:
        data = await state.get_data()
        original_message_id = data.get("original_message_id")
        await bot.edit_message_reply_markup(
            chat_id=callback.from_user.id, message_id=original_message_id, reply_markup=None)
    except Exception:
        pass
    await state.clear()
    await callback.message.answer('Отменено', reply_markup=admin_menu)


@admin_router.message(F.text == 'Статистика по объектам')
async def start_house(message: Message, state: FSMContext, session: AsyncSession):
    """
    Обработчик команды 'Статистика по объектам'.
    Загружает список объектов из БД и начинает показ с первой карточки.
    """
    houses = await orm_get_start_12(session)  # Получаем список объектов недвижимости из БД
    if not houses:
        await message.answer("Объявлений пока нет.")  # Если нет объектов — уведомляем пользователя
        return

    print(houses[0].id)  # Отладочный вывод ID первого объекта
    await state.update_data(houses=houses, index=0)  # Сохраняем список и индекс в состояние FSM
    await send_house_card1(message, houses[0], 0, state, session)  # Отправляем карточку с первым объектом


async def send_house_card1(message, house, index, state: FSMContext, session: AsyncSession):
    """
    Отправляет карточку с информацией об объекте недвижимости.
    Включает навигационные кнопки для перехода к следующему/предыдущему объекту.
    """
    data = await state.get_data()
    # print(data)

    houses = data.get("houses", [])
    status_item = await orm_get_start_id_status(session, house.id)  # Получаем статус объекта (лайки, просмотры)

    if house.status == 'продан':
        text = (f"<b>ПРОДАН</b>\n\n"
                f"{house.fio_buy}\n\n"
                f"Вид объекта: {house.title}\n"
            f"Местоположение: {house.address}\n"

            f"Ремонт: {house.repair}\n"
            f"Площадь дома м²: {house.area}\n"
            f"Тип здания: {house.building_type}\n"
            f"Цена: {house.price}\n\n"
            f"Природный газ: {house.gas_supply}\n\n"
                f"Просмотров: {status_item.get('watch')}\n"
                f"Лайков: {status_item.get('like')}")
    else:
        text = (f"Вид объекта: {house.title}\n"
            f"Местоположение: {house.address}\n"

            f"Ремонт: {house.repair}\n"
            f"Площадь дома м²: {house.area}\n"
            f"Тип здания: {house.building_type}\n"
            f"Цена: {house.price}\n\n"
            f"Природный газ: {house.gas_supply}\n\n"
                f"Просмотров: {status_item.get('watch')}\n"
                f"Лайков: {status_item.get('like')}")

    kb_row = []

    # Добавляем кнопку "назад" если это не первый объект
    if index > 0:
        kb_row.append(InlineKeyboardButton(text="◀️", callback_data="prev_house1"))

    # Добавляем кнопку "вперёд" если это не последний объект
    if index < len(houses) - 1:
        kb_row.append(InlineKeyboardButton(text="▶️", callback_data="next_house1"))

    # Формируем клавиатуру с навигацией и кнопкой отмены
    kb = InlineKeyboardMarkup(inline_keyboard=[
        kb_row,
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_houses1")]
    ])

    # Отправляем фото с подписью и клавиатурой
    try:
        if f'{house.photo}'.startswith('/BOT'):
            await message.answer_photo(photo=FSInputFile(house.photo), caption=text, reply_markup=kb)
        else:
            await message.answer_photo(photo=house.photo, caption=text, reply_markup=kb)
    except:
        await message.answer(text=text, reply_markup=kb)


@user_router.callback_query(F.data.in_({'next_house1', 'prev_house1'}))
async def paginate_houses(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    """
    Обработка нажатий кнопок навигации вперед/назад в просмотре объектов.
    Изменяет индекс текущего объекта и обновляет карточку.
    """
    data = await state.get_data()
    houses = data.get('houses')
    index = data.get('index', 0)

    # Переключаем индекс в зависимости от кнопки
    if callback.data == 'next_house1':
        index = (index + 1) % len(houses)
    else:
        index = (index - 1) % len(houses)

    await state.update_data(index=index)  # Обновляем индекс в состоянии
    await callback.message.delete()  # Удаляем старое сообщение с карточкой
    await send_house_card1(callback.message, houses[index], index, state, session)  # Отправляем новую карточку


@user_router.callback_query(F.data == 'cancel_houses1')
async def cancel_house_view(callback: CallbackQuery, state: FSMContext):
    """
    Обработка нажатия кнопки отмены просмотра объявлений.
    Очищает состояние FSM и удаляет сообщение.
    """
    await state.clear()  # Очищаем данные состояния
    await callback.message.delete()  # Удаляем сообщение с карточкой
    await callback.message.answer("Просмотр объявлений завершён.", reply_markup=None)


# --- Добавление администратора ---
class Admin(StatesGroup):
    step_1 = State()
    step_2 = State()


@admin_router.message(F.text == 'Добавить администратора')
async def start_house(message: Message, bot: Bot, state: FSMContext):
    """
    Запускает процесс добавления администратора.
    Запрашивает у пользователя ID нового администратора.
    """
    await message.answer('Введите id кого добавить админом')
    await state.set_state(Admin.step_1)


@admin_router.message(Admin.step_1, F.text)
async def mes_admin(message: Message, state: FSMContext, bot: Bot):
    """
    Получает ID нового администратора и добавляет в список.
    """
    admin_list = [int(message.text)]
    bot.my_admins_list += admin_list
    await message.answer('Админ добавлен')
    await state.clear()


# --- Вывод всех объектов в Excel ---
@admin_router.message(F.text == 'Вывести все объекты в excel')
async def send_file_review(message: Message):
    """
    Формирует Excel-файл со всеми объектами недвижимости из БД
    и отправляет его пользователю.
    """
    conn = sqlite3.connect('my_base.db')
    query = "SELECT * FROM registration_house"
    df = pd.read_sql_query(query, conn)
    conn.close()

    # Переименование колонок для читаемости в Excel
    df.rename(columns={
        'id': 'ID',
        'title': 'Тип объекта',
        'address': 'Местоположение',
        'repair': 'Ремонт',
        'area': 'Площадь дома м²',
        'building_type': 'Тип здания',
        'price': 'Цена',
        'gas_supply': 'Природный газ',
        'status': 'Статус',
        # 'photo': 'Фото (ссылки)',
    }, inplace=True)

    excel_path = 'Объекты.xlsx'
    df.to_excel(excel_path, index=False)

    # Подгонка ширины колонок под содержимое
    wb = load_workbook(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 8
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_path)

    # Отправка файла пользователю
    await message.reply_document(document=FSInputFile(path=excel_path))


# --- Аналогично для статистики ---
@admin_router.message(F.text == 'Статистика по объектам в excel')
async def send_file_review(message: Message):
    """
    Формирует Excel-файл со статистикой по объектам (лайки, просмотры)
    и отправляет его пользователю.
    """
    conn = sqlite3.connect('my_base.db')
    query = "SELECT * FROM registration_status"
    df = pd.read_sql_query(query, conn)
    conn.close()

    df.rename(columns={
        'id': 'ID',
        'id_house': 'ID строения',
        'like': 'поставлено лайков',
        'watch': 'просмотрено',
    }, inplace=True)

    excel_path = 'Статистика по объектам.xlsx'
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2  # отступ
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_path)

    await message.reply_document(document=FSInputFile(path=excel_path))


# --- Изменение статуса объекта на "продан" ---
class Status(StatesGroup):
    step_1 = State()
    step_2 = State()
    step_3 = State()


@admin_router.message(F.text == 'Изменить статус объекта на продан')
async def send_file_review(message: Message, state: FSMContext):
    """
    Начинает процесс изменения статуса объекта.
    Отправляет файл с объектами и просит ввести ID объекта.
    """
    conn = sqlite3.connect('my_base.db')
    query = "SELECT * FROM registration_house"
    df = pd.read_sql_query(query, conn)
    conn.close()

    df.rename(columns={
        'id': 'ID',
        'title': 'Тип объекта',
        'address': 'Местоположение',
        'repair': 'Ремонт',
        'area': 'Площадь дома м²',
        'building_type': 'Тип здания',
        'price': 'Цена',
        'gas_supply': 'Природный газ',
        'status': 'Статус',
        # 'photo': 'Фото (ссылки)',
    }, inplace=True)

    excel_path = 'Объекты.xlsx'
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 8  # отступ
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_path)

    await message.reply_document(document=FSInputFile(path=excel_path),
                                 caption='Введите id объекта которого изменим статус')
    await state.set_state(Status.step_1)


@admin_router.message(Status.step_1, F.text)
async def status_mess(message: Message, state: FSMContext, session: AsyncSession):
    """
    Получает ID объекта от пользователя и меняет его статус на 'продан'.
    """
    dta = await orm_get_start_id_life(session, int(message.text))
    if dta is None:
        await message.answer('Ошибка, нет такого объекта\nВведите id объекта которого изменим статус')
    else:
        await state.update_data(id_object=message.text)
        await message.answer('Видите id человека кто купил!')
        await state.set_state(Status.step_2)


@admin_router.message(Status.step_2, F.text)
async def status_mess(message: Message, state: FSMContext, session: AsyncSession):
    data_id = await orm_get_start_id_people(session, int(message.text))
    if data_id is None:
        await message.answer('Этого пользователя нету в базе, Введите его ФИО')
        await state.set_state(Status.step_3)
    else:
        data = await state.get_data()
        text = f'{data_id.get("fio")}\n{data_id.get("phone")}'
        await orm_set_status_sold(session, int(data.get('id_object')), text)
        await message.answer('Статус изменен', reply_markup=admin_menu)
        await state.clear()


@admin_router.message(Status.step_3, F.text)
async def status_mess(message: Message, state: FSMContext, session: AsyncSession):
    data = await state.get_data()
    await orm_set_status_sold(session, int(data.get('id_object')), message.text)
    await message.answer('Статус изменен', reply_markup=admin_menu)
    await state.clear()


class File(StatesGroup):
    step_1 = State()


@admin_router.message(F.text == 'Добавить объекты в excel')
async def status_mess(message: Message, state: FSMContext, ):
    await message.answer('Загрузите объект')
    await state.set_state(File.step_1)


@user_router.message(File.step_1, F.document)
async def handle_excel_file(message: Message, bot: Bot, session: AsyncSession, state: FSMContext):
    document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.reply("Пожалуйста, загрузите файл в формате .xlsx")
        return

    file_path = f"downloads/{document.file_name}"
    os.makedirs("downloads", exist_ok=True)

    file = await bot.get_file(document.file_id)
    await bot.download_file(file.file_path, destination=file_path)

    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        # Пропускаем первую строку и собираем остальные
        data_rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # Пропускаем заголовки
            # if any(cell is None for cell in row):
            #     await message.answer(f"Ошибка при обработке строки: {i}")
            # else:
            for cell in row:
                if isinstance(cell, str) and '|' in cell:
                    # Делим строку на элементы
                    parts = [part.strip() for part in cell.split('|')]
                    data_rows.extend(parts)
                else:
                    data_rows.append(cell)
            data = {
                'title': data_rows[1],
                'address': data_rows[3],
                'repair': data_rows[4],
                'area': data_rows[5],
                'building_type': data_rows[6],
                'price': data_rows[2],
                'gas_supply': data_rows[7],
                'photo': data_rows[8],

                'status': 'Продается',
                'fio_buy': None
            }
            id_status = await orm_add_house(session, data)
            status = {
                'id_house': id_status,
                'like': 0,
                'like_nick': None,
                'watch': 0,
            }
            await orm_add_status(session, status)
            # await message.answer(f'{i} успешно опубликовано')
            data_rows = []
        await message.answer(f'успешно опубликовано')

    except Exception as e:
        await message.answer(f"Ошибка при обработке файла: {e}")
    finally:
        os.remove(file_path)
        await state.clear()
