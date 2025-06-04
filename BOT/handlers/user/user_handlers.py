import sqlite3

import pandas as pd
from aiogram import Bot, Router, F
from dotenv import load_dotenv

from aiogram.filters import Command, or_f
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove, \
    InputFile, FSInputFile
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

from database.orm_database import orm_get_start, orm_increment_status, orm_add_people, \
    add_to_like_people, orm_get_start_id_people, orm_append_to_status, orm_get_start_id_life, orm_get_filtered
from kb.replay import phone_count
from kb.user_inline import inline_price_house, inline_price_house_filtr

load_dotenv()

user_router = Router()


class People(StatesGroup):
    step_1 = State()


class PeopleRegistr(StatesGroup):
    step_1 = State()
    step_2 = State()


@user_router.message(Command('start'))
async def start_command(message: Message, session: AsyncSession, state: FSMContext):
    """
    Обработчик команды /start.

    Добавляет пользователя в базу (если он ещё не добавлен),
    после чего отправляет приветственное сообщение с клавиатурой выбора действия.

    Args:
        message (Message): Объект сообщения от пользователя.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
    """
    yes_people = await orm_get_start_id_people(session, message.from_user.id)
    if yes_people is None:
        await message.answer('Добрый день, Вас приветствует небольшая регистрация,\n'
                             'введите Ваше ФИО (пример: Иванов Иван Иванович)')
        await state.set_state(PeopleRegistr.step_1)
    else:

        await message.answer('Выберите действие', reply_markup=await inline_price_house())


@user_router.message(PeopleRegistr.step_1, F.text)
async def regist_people(message: Message, state: FSMContext):
    await state.update_data(fio=message.text)
    await message.answer('Теперь подтвердите номер телефона или видите вручную', reply_markup=phone_count)
    await state.set_state(PeopleRegistr.step_2)


@user_router.message(PeopleRegistr.step_2, or_f(F.text, F.contact))
async def phone_contact(message: Message, state: FSMContext, session: AsyncSession):
    if message.contact:
        await state.update_data(step_7=message.contact.phone_number)
    else:
        await state.update_data(step_7=message.text)
    data_fio = await state.get_data()
    data = {
        'id_people': message.from_user.id,
        'nick': message.from_user.username,
        'fio': data_fio.get('fio'),
        'phone': data_fio.get('step_7'),
        'like_status': None,
    }
    await orm_add_people(session, data)
    await message.answer('Спасибо за регистрацию', reply_markup=ReplyKeyboardRemove())
    await message.answer('Выберите действие', reply_markup=await inline_price_house())
    await state.clear()


@user_router.callback_query(F.data == 'показать_объявления')
async def start_house(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    """
    Обработчик нажатия кнопки "показать_объявления".

    Получает список всех объявлений из базы,
    если они есть — сохраняет их и индекс текущего объявления в состоянии,
    отправляет карточку первого объявления.

    Args:
        callback (CallbackQuery): Объект коллбэка от нажатия кнопки.
        state (FSMContext): Контекст состояния пользователя.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
    """
    houses = await orm_get_start(session)
    if not houses:
        await callback.message.answer("Объявлений пока нет.")
        return

    print(houses[0].id)
    await state.update_data(houses=houses, index=0)
    await send_house_card(callback.message, houses[0], 0, state, session)


async def send_house_card(message, house, index, state: FSMContext, session: AsyncSession):
    """
    Отправляет карточку с информацией об объявлении и клавиатуру навигации.

    Формирует текст с данными объекта недвижимости и кнопки "вперед", "назад", "лайк" и "отмена".
    При отправке карточки увеличивает счетчик просмотров объявления.

    Args:
        message (Message): Сообщение, к которому будет прикреплена карточка.
        house (Object): Объект объявления с необходимыми атрибутами.
        index (int): Индекс текущего объявления в списке.
        state (FSMContext): Контекст состояния пользователя.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
    """
    data = await state.get_data()
    houses = data.get("houses", [])

    text = (f"Вид объекта: {house.title}\n"
            f"Местоположение: {house.address}\n"

            f"Ремонт: {house.repair}\n"
            f"Площадь дома м²: {house.area}\n"
            f"Тип здания: {house.building_type}\n"
            f"Цена: {house.price}\n\n"
            f"Природный газ: {house.gas_supply}\n")

    kb_row = []

    if index > 0:
        kb_row.append(InlineKeyboardButton(text="◀️", callback_data="prev_house"))

    kb_row.append(InlineKeyboardButton(text="❤️", callback_data=f"like_{index}"))

    if index < len(houses) - 1:
        kb_row.append(InlineKeyboardButton(text="▶️", callback_data="next_house"))

    kb = InlineKeyboardMarkup(inline_keyboard=[
        kb_row, [InlineKeyboardButton(text="Попросить связаться", callback_data=f"tap_house-{houses[index].id}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_houses")]
    ])

    # Увеличиваем счетчик просмотров у объявления
    await orm_increment_status(session, product_id=int(houses[index].id), field="watch")
    try:
        if f'{house.photo}'.startswith('/BOT'):
            await message.answer_photo(photo=FSInputFile(house.photo), caption=text, reply_markup=kb)
        else:
            await message.answer_photo(photo=house.photo, caption=text, reply_markup=kb)
    except:
        await message.answer(text=text, reply_markup=kb)


@user_router.callback_query(F.data.startswith('tap_house'))
async def paginate_houses(callback: CallbackQuery, state: FSMContext, session: AsyncSession, bot: Bot):
    await callback.answer('Отлично. Ожидайте', show_alert=True)
    index = int(callback.data.split('-')[1])
    people_list = await orm_get_start_id_people(session, callback.from_user.id)
    id_house = await orm_get_start_id_life(session, index)

    # houses = data.get("houses", [])
    text = (f"НИК: @{people_list.get('nick')}\n"
            f"ФИО: @{people_list.get('fio')}\n"
            f"Телефон: @{people_list.get('phone')}\n\n"
            f"Вид объекта: {id_house.get('title')}\n"
            f"Местоположение: {id_house.get('address')}\n"
            f"Ремонт: {id_house.get('repair')}\n"
            f"Площадь дома м²: {id_house.get('area')}\n"
            f"Тип здания: {id_house.get('building_type')}\n"
            f"Цена: {id_house.get('price')}\n\n"
            f"Природный газ: {id_house.get('gas_supply')}\n")
    try:
        await bot.send_photo(-1002489940696, photo=id_house.get('photo'), caption=text)
    except:
        await bot.send_message(-1002489940696, text)


@user_router.callback_query(F.data.in_({'next_house', 'prev_house'}))
async def paginate_houses(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    """
    Обработчик пагинации объявлений.

    Обрабатывает нажатия кнопок "вперед" и "назад",
    меняет индекс текущего объявления, удаляет предыдущее сообщение
    и отправляет карточку с новым объявлением.

    Args:
        callback (CallbackQuery): Объект коллбэка от нажатия кнопки.
        state (FSMContext): Контекст состояния пользователя.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
    """
    data = await state.get_data()
    houses = data.get('houses')
    index = data.get('index', 0)

    if callback.data == 'next_house':
        index = (index + 1) % len(houses)
    else:
        index = (index - 1) % len(houses)

    await state.update_data(index=index)
    await callback.message.delete()
    await send_house_card(callback.message, houses[index], index, state, session)


@user_router.callback_query(F.data == 'cancel_houses')
async def cancel_house_view(callback: CallbackQuery, state: FSMContext):
    """
    Обработчик отмены просмотра объявлений.

    Очищает состояние, удаляет сообщение с карточкой,
    отправляет сообщение о завершении просмотра с клавиатурой.

    Args:
        callback (CallbackQuery): Объект коллбэка от нажатия кнопки.
        state (FSMContext): Контекст состояния пользователя.
    """
    await state.clear()
    await callback.message.delete()
    await callback.message.answer("Просмотр объявлений завершён.", reply_markup=await inline_price_house())


@user_router.callback_query(F.data.startswith('like_'))
async def like_house(callback: CallbackQuery, session: AsyncSession, state: FSMContext):
    """
    Обработчик добавления объявления в избранное.

    Проверяет, добавлено ли объявление пользователем в избранное,
    если нет — добавляет в базу, увеличивает счетчик лайков,
    добавляет пользователя в список лайкнувших.

    Args:
        callback (CallbackQuery): Объект коллбэка от нажатия кнопки.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
        state (FSMContext): Контекст состояния пользователя.
    """
    index = int(callback.data.split('_')[1])
    data = await state.get_data()
    houses = data.get("houses", [])
    people_list = await orm_get_start_id_people(session, callback.from_user.id)

    # Проверяем, лайкнул ли уже пользователь это объявление
    if isinstance(people_list["like_status"], dict) and f"{houses[index].id}" in people_list["like_status"]:
        await callback.answer(f"Объявление уже добавлено в избранное ❤️", show_alert=True)
    else:
        await add_to_like_people(session, callback.from_user.id, key=houses[index].id, value=1)
        await orm_increment_status(session, product_id=houses[index].id, field="like")
        await orm_append_to_status(session, houses[index].id, callback.from_user.username)
        await callback.answer(f"Объявление №{index + 1} добавлено в избранное ❤️", show_alert=True)


@user_router.callback_query(F.data == 'показать_избранные_объявления')
async def start_house(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    """
    Обработчик показа избранных объявлений пользователя.

    Получает список избранных объявлений пользователя,
    фильтрует все объявления, оставляя только избранные,
    отправляет первую карточку из этого списка.

    Args:
        callback (CallbackQuery): Объект коллбэка от нажатия кнопки.
        state (FSMContext): Контекст состояния пользователя.
        session (AsyncSession): Асинхронная сессия работы с базой данных.
    """
    people_list = await orm_get_start_id_people(session, callback.from_user.id)
    if people_list["like_status"] is None:
        await callback.message.answer("Объявлений пока нет.")
    else:
        houses = await orm_get_start(session)
        cor_item = []
        final_item = []
        # Собираем ID избранных объявлений
        for key in people_list["like_status"]:
            cor_item.append(int(key))
        # Фильтруем объявления по избранным ID
        for i in houses:
            if i.id in cor_item:
                final_item.append(i)
        await state.update_data(houses=final_item, index=0)
        await send_house_card(callback.message, final_item[0], 0, state, session)


@user_router.callback_query(F.data == 'exel')
async def send_file_review(callback: CallbackQuery):
    await callback.answer()
    """
    Формирует Excel-файл со всеми объектами недвижимости из БД,
    исключая проданные, и отправляет его пользователю.
    """
    conn = sqlite3.connect('my_base.db')
    query = "SELECT * FROM registration_house"
    df = pd.read_sql_query(query, conn)
    conn.close()

    # Удаляем объекты со статусом 'продан' (без учёта регистра и лишних пробелов)
    df = df[~df['status'].str.strip().str.lower().eq('продан')]
    df.drop(columns=['id'], inplace=True)
    df.drop(columns=['photo'], inplace=True)
    df.drop(columns=['fio_buy'], inplace=True)

    df.insert(0, '№', range(1, len(df) + 1))

    # Переименование колонок для читаемости в Excel
    df.rename(columns={
        'title': 'Тип объекта',
        'address': 'Местоположение',
        'repair': 'Ремонт',
        'area': 'Площадь дома м²',
        'building_type': 'Тип здания',
        'price': 'Цена',
        'gas_supply': 'Природный газ',
        'status': 'Статус',
        # 'photo': 'Фото (ссылки)',
    }, inplace=True)

    excel_path = 'Объекты.xlsx'
    df.to_excel(excel_path, index=False)

    # Подгонка ширины колонок под содержимое
    wb = load_workbook(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 8
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_path)

    # Отправка файла пользователю
    await callback.message.reply_document(document=FSInputFile(path=excel_path))



@user_router.callback_query(F.data == 'фильтр_объявлений')
async def filter_menu(callback: CallbackQuery, state: FSMContext):
    """
    Показывает пользователю фильтры для выбора.
    """

    await state.clear()
    await callback.message.edit_text("Выберите фильтры:", reply_markup=await inline_price_house_filtr())


@user_router.callback_query(F.data.startswith('filter_category_'))
async def set_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data.replace("filter_category_", "")
    await state.update_data(filter_category=category)
    await callback.answer(f"Выбрана категория: {category}")


@user_router.callback_query(F.data.startswith('filter_price_'))
async def set_price_filter(callback: CallbackQuery, state: FSMContext):
    max_price = int(callback.data.replace("filter_price_", ""))
    await state.update_data(filter_price=max_price)
    await callback.answer(f"Цена до {max_price}₽ выбрана")


@user_router.callback_query(F.data == 'показать_отфильтрованные')
async def show_filtered(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    data = await state.get_data()
    category = data.get("filter_category")
    max_price = data.get("filter_price")  # Добавляем цену

    # Можешь напечатать для отладки:
    print("Категория из FSM:", category)

    houses = await orm_get_filtered(session, category=category, max_price=max_price)
    if not houses:
        await callback.message.answer("Нет объявлений по выбранному фильтру.")
        return

    await state.update_data(houses=houses, index=0)
    await send_house_card(callback.message, houses[0], 0, state, session)



